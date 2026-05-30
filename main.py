from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import requests
import psycopg2
import psycopg2.extras
import os
import json
import re
import io
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = FastAPI(title="Refinery Contract Chatbot API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Static files & root route
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def root():
    return FileResponse("static/index.html")

# -- Config -------------------------------------------------------------------
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://postgres:password@host:5432/railway")
DINOIKI_API_KEY = os.getenv("DINOIKI_API_KEY", "")
DINOIKI_URL = "https://ai.dinoiki.com/v1/chat/completions"
AI_MODEL = "gpt-4o"

# -- Stopwords bahasa Indonesia ------------------------------------------------
STOPWORDS = {
    'apa', 'siapa', 'berapa', 'yang', 'dan', 'atau', 'dari', 'untuk',
    'dengan', 'adalah', 'ini', 'itu', 'ada', 'tidak', 'bisa', 'mau',
    'saya', 'kamu', 'dia', 'kami', 'kita', 'mereka', 'semua', 'sudah',
    'belum', 'sedang', 'akan', 'telah', 'pada', 'di', 'ke', 'oleh',
    'juga', 'hanya', 'lebih', 'paling', 'sangat', 'banyak', 'sedikit',
    'tampilkan', 'tunjukkan', 'cari', 'lihat', 'data', 'info', 'informasi',
    'list', 'daftar', 'total', 'jumlah', 'nilai', 'status', 'semua',
    'kontrak', 'vendor', 'tagihan', 'dokumen', 'progress', 'bulan', 'tahun'
}

# -- Helper: call dinoiki AI --------------------------------------------------
def call_ai(messages: list, max_tokens: int = 1500) -> str:
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {DINOIKI_API_KEY}"
    }
    payload = {
        "model": AI_MODEL,
        "messages": messages,
        "max_tokens": max_tokens,
        "temperature": 0.3,
    }
    resp = requests.post(DINOIKI_URL, headers=headers, json=payload, timeout=60)
    resp.raise_for_status()
    data = resp.json()
    return data["choices"][0]["message"]["content"].strip()

# -- Dynamic Context Injection -------------------------------------------------
def smart_entity_search(user_message: str) -> str:
    try:
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()
        context = []
        found_ids = set()

        words = [w for w in re.findall(r'\b\w{3,}\b', user_message)
                 if w.lower() not in STOPWORDS]
        company_patterns = re.findall(r'\b(?:PT|CV|UD|TB|PD)\s+[\w\s]+', user_message, re.IGNORECASE)
        search_terms = list(set(words + company_patterns))

        for term in search_terms:
            term = term.strip()
            if len(term) < 3:
                continue

            cur.execute("""
                SELECT id_vendor, nama_vendor, status_vendor, score
                FROM vendor WHERE nama_vendor ILIKE %s LIMIT 3
            """, (f'%{term}%',))
            for v in cur.fetchall():
                key = f"vendor_{v[0]}"
                if key not in found_ids:
                    found_ids.add(key)
                    context.append(f"[VENDOR] '{v[1]}' -> id_vendor={v[0]}, status={v[2]}, score={v[3]}")

            cur.execute("""
                SELECT k.id_kontrak, k.judul_kontrak, k.no_dokumen_kontrak,
                       k.direksi_pekerjaan, k.status_kontrak, k.tipe_kontrak, v.nama_vendor
                FROM kontrak k LEFT JOIN vendor v ON k.id_vendor = v.id_vendor
                WHERE k.judul_kontrak ILIKE %s OR k.no_dokumen_kontrak ILIKE %s
                   OR k.no_po_pr ILIKE %s OR k.direksi_pekerjaan ILIKE %s LIMIT 3
            """, (f'%{term}%', f'%{term}%', f'%{term}%', f'%{term}%'))
            for k in cur.fetchall():
                key = f"kontrak_{k[0]}"
                if key not in found_ids:
                    found_ids.add(key)
                    context.append(f"[KONTRAK] '{k[1]}' -> id_kontrak={k[0]}, doc={k[2]}, direksi={k[3]}, status={k[4]}, tipe={k[5]}, vendor='{k[6]}'")

            cur.execute("""
                SELECT t.id_tagihan, t.nomor_tagihan, t.status_tagihan, t.nilai_tagihan, k.judul_kontrak
                FROM tagihan t LEFT JOIN kontrak k ON t.id_kontrak = k.id_kontrak
                WHERE t.nomor_tagihan ILIKE %s LIMIT 3
            """, (f'%{term}%',))
            for t in cur.fetchall():
                key = f"tagihan_{t[0]}"
                if key not in found_ids:
                    found_ids.add(key)
                    context.append(f"[TAGIHAN] '{t[1]}' -> id_tagihan={t[0]}, status={t[2]}, nilai={t[3]}, kontrak='{t[4]}'")

            cur.execute("""
                SELECT p.id_padi, p.no_pembelian, p.judul_pembelian, p.nilai, v.nama_vendor
                FROM padi p LEFT JOIN vendor v ON p.id_vendor = v.id_vendor
                WHERE p.no_pembelian ILIKE %s OR p.judul_pembelian ILIKE %s LIMIT 3
            """, (f'%{term}%', f'%{term}%'))
            for p in cur.fetchall():
                key = f"padi_{p[0]}"
                if key not in found_ids:
                    found_ids.add(key)
                    context.append(f"[PADI] '{p[2]}' -> id_padi={p[0]}, no_pembelian={p[1]}, nilai={p[3]}, vendor='{p[4]}'")

        conn.close()

        if context:
            result = "\n\nKONTEKS ENTITAS YANG DITEMUKAN DI DATABASE:\n"
            result += "(Gunakan informasi ini untuk memahami maksud user tanpa perlu klarifikasi)\n"
            result += "\n".join(context)
            return result

        return ""

    except Exception as e:
        return ""

# -- DB Schema context untuk AI -----------------------------------------------
SCHEMA_CONTEXT = """
Database PostgreSQL untuk sistem manajemen kontrak kilang minyak. Berikut skema tabel:

TABEL: profiles
Kolom: id, email, full_name, role (admin/pic/user), password_hash, created_at, updated_at, is_active, id_vendor

TABEL: vendor
Kolom: id_vendor, nama_vendor, npwp, alamat, pic_nama, pic_kontak, status_vendor (Active/Inactive/Blacklist), score, created_at, updated_at

TABEL: kontrak
Kolom: id_kontrak, id_vendor, judul_kontrak, no_dokumen_kontrak, no_po_pr, direksi_pekerjaan,
  tipe_kontrak (Lumpsum/Unit Price/TSA/LTSA/TSA-LTSA), status_kontrak (Pre-KOM/Active/Aktif/Completed/Selesai/Terminated),
  tanggal_spb_diterima, tanggal_terima_dokumen, tanggal_maksimal_kom, tanggal_mulai, tanggal_selesai,
  sla_kom_hari, estimasi_tanggal_kom, tanggal_kom, kom_terlambat, nilai_awal, durasi_kontrak_hari,
  progress_plan, progress_actual, aktivitas_saat_ini, kendala, disiplin, tkdn_percentage, tanggal_lkp,
  has_amendment, no_amandemen, tanggal_amandemen, jenis_amandemen, nilai_kontrak_baru, durasi_amandemen,
  tanggal_mulai_baru, tanggal_selesai_baru, alasan_perubahan, contract_documents, amendment_documents,
  s_curve_data, tanggal_mpl, tanggal_mpa, masa_pemeliharaan_hari, created_at, updated_at

TABEL: amandemen_kontrak
Kolom: id_amandemen, id_kontrak, nomor_urut, no_amandemen, tanggal_amandemen, jenis_amandemen,
  nilai_kontrak_baru, durasi_amandemen, tanggal_mulai_baru, tanggal_selesai_baru, alasan_perubahan,
  amendment_documents, created_at, updated_at

TABEL: tagihan
Kolom: id_tagihan, id_kontrak, nomor_tagihan, tanggal_tagihan, tipe_kontrak, termin, nilai_tagihan,
  status_tagihan, memo_required, tanggal_pengiriman_memo, dokumen_memo, dokumen_tagihan, catatan,
  created_at, updated_at

TABEL: progress_lumpsum
Kolom: id_progress, id_kontrak, milestone, persen, tanggal_update, evidence, created_at

TABEL: progress_unit_price
Kolom: id_progress, id_kontrak, nama_item, satuan, qty_rencana, qty_aktual, harga_satuan, tanggal_update, created_at

TABEL: monitoring_ltsa
Kolom: id_log, id_kontrak, tanggal_kunjungan, jenis_layanan (Preventive/Corrective/Standby),
  durasi_jam, sla_terpenuhi (Yes/No), keterangan, created_at

TABEL: padi
Kolom: id_padi, no_pembelian, tanggal, judul_pembelian, no_po_pr, nilai, id_vendor, link_pembelian,
  bagian, dokumen_pendukung, status_purchase (BAST), tanggal_bast, tanggal_sa_gr, tanggal_invoice,
  tanggal_payment_approval, tanggal_paid, catatan_status, created_at, updated_at

TABEL: dokumen_approval
Kolom: id_dokumen, id_kontrak, tipe_dokumen (Evident/Report/Persetujuan), nama_dokumen,
  deskripsi_dokumen, file_path, file_url, nama_file, tipe_file, ukuran_file,
  status_approval (Pending/Approved/Rejected), catatan_reviewer, uploaded_by, reviewed_by,
  reviewed_at, created_at, updated_at

TABEL: konfigurasi_sistem
Kolom: id_setting, nama_setting, nilai_setting, deskripsi, updated_at

TABEL: daily_report
Kolom: id_report, tanggal_laporan, disiplin (Electrical/Instrument/Rotating/Stationary/Alat Berat),
  direksi (MA5/MA6/MA7/Workshop), kategori (Corrective Maintenance/Preventive Maintenance/Plant Patrol/Progress/Challenge Session),
  tag_number, deskripsi, status_pekerjaan (Done/In Progress/Waiting Material/Pending/-),
  catatan, pengirim_wa, raw_text, created_at

Relasi penting:
- vendor.id_vendor -> kontrak.id_vendor (1 vendor banyak kontrak)
- kontrak.id_kontrak -> tagihan.id_kontrak
- kontrak.id_kontrak -> amandemen_kontrak.id_kontrak
- kontrak.id_kontrak -> progress_lumpsum.id_kontrak
- kontrak.id_kontrak -> progress_unit_price.id_kontrak
- kontrak.id_kontrak -> monitoring_ltsa.id_kontrak
- kontrak.id_kontrak -> dokumen_approval.id_kontrak
- vendor.id_vendor -> padi.id_vendor

NILAI ENUM & PILIHAN YANG VALID:

1. TIPE KONTRAK: 'Lumpsum', 'Unit Price', 'TSA', 'LTSA', 'TSA/LTSA'
2. STATUS KONTRAK: 'Pre-KOM', 'Aktif', 'Selesai', 'Terminated'
3. DISIPLIN: 'Instrumentasi', 'Stationary', 'Electrical', 'Rotating', 'Alat Berat'
4. DIREKSI PEKERJAAN: 'MA5', 'MA6', 'MA7', 'Workshop'
5. JENIS AMANDEMEN: 'Nilai', 'Waktu', 'Nilai dan Waktu'
6. STATUS APPROVAL: 'Pending', 'Approved', 'Rejected'
7. STATUS VENDOR: 'Active', 'Inactive', 'Blacklist'
8. JENIS LAYANAN LTSA: 'Preventive', 'Corrective', 'Standby'
9. STATUS TAGIHAN (urutan tahapan):
   LKP -> Punchlist -> BAST -> BAKP/BAPP -> Submit i-Vendor -> SA -> PA -> Verification -> Payment/Selesai
10. STATUS PURCHASE PADI: 'BAST'
"""

BASE_SYSTEM_PROMPT = (
    "Kamu adalah asisten cerdas untuk sistem manajemen kontrak kilang minyak.\n"
    "Kamu dapat menjawab pertanyaan bisnis dalam bahasa Indonesia secara natural "
    "dan mengkonversinya ke query SQL PostgreSQL.\n\n"
    + SCHEMA_CONTEXT +
    "\nATURAN KETAT:\n"
    "1. HANYA boleh generate query SELECT, TIDAK boleh UPDATE, DELETE, INSERT, DROP, ALTER, TRUNCATE, dll\n"
    "2. TIDAK boleh query SELECT * (tanpa kolom spesifik) - selalu tentukan kolom yang relevan\n"
    "3. Selalu gunakan LIMIT maksimal 1000 baris\n"
    "4. Gunakan JOIN yang tepat antar tabel\n"
    "5. Format angka nilai kontrak dalam format Indonesia (Rp)\n"
    "\nATURAN INTERPRETASI ENTITAS:\n"
    '- Jika ada blok "KONTEKS ENTITAS YANG DITEMUKAN DI DATABASE" -> gunakan langsung, JANGAN minta klarifikasi\n'
    "- Jika user menyebut nama yang diawali PT/CV/UD -> cari di vendor.nama_vendor\n"
    "- Jika user menyebut kode seperti MA5, KOM-001 -> cari di direksi_pekerjaan atau no_dokumen_kontrak\n"
    "- Jika entitas tidak ditemukan di konteks -> baru boleh minta klarifikasi\n"
    "\nFORMAT RESPONS JSON:\n"
    "Kamu HARUS selalu merespons dalam format JSON seperti ini:\n"
    "{\n"
    '  "type": "query" | "clarification" | "narrative" | "error",\n'
    '  "sql": "query SQL jika type=query",\n'
    '  "explanation": "penjelasan dalam bahasa Indonesia apa yang akan dilakukan query ini",\n'
    '  "narrative_hint": "bagaimana cara menarasikan hasilnya nanti",\n'
    '  "chart_suggestion": null | "bar" | "line" | "pie" | "doughnut",\n'
    '  "chart_config": null | {"x_column": "...", "y_column": "...", "label": "..."},\n'
    '  "clarification_question": "pertanyaan klarifikasi jika type=clarification",\n'
    '  "message": "pesan untuk user"\n'
    "}\n"
    "\nDETEKSI CHART:\n"
    '- Jika pertanyaan menyebut "grafik", "chart", "trend", "perbandingan", "distribusi", "per bulan/tahun" -> suggest chart\n'
    "- bar chart: perbandingan kategori\n"
    "- line chart: data time-series\n"
    "- pie/doughnut: distribusi persentase\n"
)

# -- Laporan System Prompt -----------------------------------------------------
LAPORAN_SYSTEM_PROMPT = (
    "Kamu adalah parser laporan harian maintenance kilang minyak.\n"
    "Tugasmu mengekstrak data dari teks laporan narasi ke dalam format JSON terstruktur.\n\n"
    "DISIPLIN YANG VALID: Electrical, Instrument, Rotating, Stationary, Alat Berat\n\n"
    "KATEGORI YANG VALID:\n"
    "- Corrective Maintenance\n"
    "- Preventive Maintenance\n"
    "- Plant Patrol\n"
    "- Progress\n"
    "- Challenge Session\n\n"
    "STATUS YANG VALID: Done, In Progress, Waiting Material, Pending, -\n\n"
    "DIREKSI (area kerja, sama dengan Bagian) YANG VALID: MA5, MA6, MA7, Workshop\n"
    "Normalisasi: 'Maintenance Area 7' / 'Area 7' / 'MA 7' / 'Bagian 7' → 'MA7'\n"
    "             'Maintenance Area 5' / 'Area 5' / 'MA 5' / 'Bagian 5' → 'MA5'\n"
    "             'Maintenance Area 6' / 'Area 6' / 'MA 6' / 'Bagian 6' → 'MA6'\n"
    "             'Workshop' → 'Workshop'\n"
    "Jika tidak ada informasi direksi, gunakan string kosong.\n\n"
    "TAG NUMBER: Kode identifikasi equipment/alat yang biasanya ada di awal deskripsi item,\n"
    "dipisah dengan titik dua (:) atau spasi. Contoh: 101-P-105, 104-P-107, 101A514.\n"
    "Jika tidak ada tag number, gunakan string kosong.\n\n"
    "ATURAN EKSTRAKSI:\n"
    "1. Satu item pekerjaan = satu entri JSON\n"
    "2. Deteksi tanggal dari teks laporan\n"
    "3. Deteksi disiplin dari header laporan\n"
    "4. Deteksi direksi dari header laporan, normalisasi ke MA5/MA6/MA7/Workshop\n"
    "5. Petakan setiap item ke kategori yang sesuai\n"
    "6. Ekstrak status dari keterangan, jika tidak ada gunakan -\n"
    "7. Ekstrak tag number dari awal deskripsi item jika ada\n"
    "8. Deskripsi diisi tanpa tag number\n\n"
    "RESPONSE FORMAT — kembalikan HANYA array JSON, tanpa teks lain:\n"
    '[\n  {\n    "tanggal_laporan": "2026-05-26",\n    "disiplin": "Instrument",\n'
    '    "direksi": "MA7",\n    "kategori": "Plant Patrol",\n    "tag_number": "105-FV-020",\n'
    '    "deskripsi": "Plant Patrol control valve",\n'
    '    "status_pekerjaan": "Done",\n    "catatan": ""\n  }\n]\n\n'
    "PENTING: Kembalikan HANYA array JSON yang valid. Jangan tambahkan penjelasan apapun."
)

# -- Laporan Functions ---------------------------------------------------------
def parse_laporan_with_ai(raw_text: str) -> list:
    try:
        response = call_ai([
            {"role": "system", "content": LAPORAN_SYSTEM_PROMPT},
            {"role": "user",   "content": f"Parse laporan berikut:\n\n{raw_text}"}
        ], max_tokens=2000)

        json_match = re.search(r'\[[\s\S]*\]', response)
        if not json_match:
            return []

        parsed = json.loads(json_match.group())
        return parsed if isinstance(parsed, list) else []

    except Exception as e:
        print(f"[PARSE LAPORAN ERROR] {e}")
        return []

def insert_daily_report(items: list, pengirim: str, raw_text: str) -> tuple:
    if not items:
        return 0, "Tidak ada item yang bisa diparse"
    try:
        conn = psycopg2.connect(DATABASE_URL)
        cur  = conn.cursor()
        success = 0
        for item in items:
            if not item.get("tanggal_laporan") or not item.get("disiplin") or not item.get("deskripsi"):
                continue
            cur.execute("""
                INSERT INTO daily_report
                    (tanggal_laporan, disiplin, direksi, kategori, tag_number, deskripsi,
                     status_pekerjaan, catatan, pengirim_wa, raw_text)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                item.get("tanggal_laporan"),
                item.get("disiplin", "-"),
                item.get("direksi", ""),
                item.get("kategori", "-"),
                item.get("tag_number", ""),
                item.get("deskripsi", "-"),
                item.get("status_pekerjaan", "-"),
                item.get("catatan", ""),
                pengirim,
                raw_text
            ))
            success += 1
        conn.commit()
        conn.close()
        return success, None
    except Exception as e:
        print(f"[INSERT LAPORAN ERROR] {e}")
        return 0, str(e)

# -- Models --------------------------------------------------------------------
class ChatRequest(BaseModel):
    message: str
    history: list = []
    pengirim: str = "web"

class LaporanRequest(BaseModel):
    raw_text: str
    pengirim: str = "web"

class DownloadRequest(BaseModel):
    sql: str
    filename: str = "data_export"

# -- DB Connection -------------------------------------------------------------
def get_db_connection():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database connection failed: {str(e)}")

# -- SQL Validator -------------------------------------------------------------
def validate_sql(sql: str) -> tuple:
    sql_upper = sql.upper().strip()

    dangerous = ["UPDATE", "DELETE", "INSERT", "DROP", "ALTER", "TRUNCATE", "CREATE", "GRANT", "REVOKE"]
    for op in dangerous:
        if re.search(r'\b' + op + r'\b', sql_upper):
            return False, f"Operasi {op} tidak diizinkan."

    if re.search(r'SELECT\s+\*', sql_upper):
        return False, "Query SELECT * tidak diizinkan."

    if not re.search(r'\bSELECT\b', sql_upper):
        return False, "Hanya query SELECT yang diizinkan."

    if "LIMIT" not in sql_upper:
        sql = sql.rstrip(";") + " LIMIT 1000"

    return True, sql

# -- Execute Query -------------------------------------------------------------
def execute_query(sql: str) -> tuple:
    valid, result = validate_sql(sql)
    if not valid:
        raise HTTPException(status_code=400, detail=result)

    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(result)
            rows    = cur.fetchall()
            columns = [desc[0] for desc in cur.description] if cur.description else []
            data    = [dict(row) for row in rows]
            return data, columns
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Query error: {str(e)}")
    finally:
        conn.close()

# -- Generate Narrative --------------------------------------------------------
def generate_narrative(data: list, columns: list, original_question: str, narrative_hint: str) -> str:
    if not data:
        return "Tidak ditemukan data yang sesuai dengan pertanyaan Anda."

    if len(data) <= 5 and len(columns) <= 5:
        data_str = json.dumps(data, default=str, ensure_ascii=False)
        try:
            return call_ai([
                {
                    "role": "system",
                    "content": "Kamu adalah asisten laporan bisnis. Jawab hanya dalam bahasa Indonesia yang profesional dan natural."
                },
                {
                    "role": "user",
                    "content": (
                        f'Pertanyaan user: "{original_question}"\n'
                        f"Hint narasi: {narrative_hint}\n"
                        f"Data hasil query: {data_str}\n\n"
                        "Buatkan narasi singkat dalam bahasa Indonesia. Maksimal 3 kalimat."
                    )
                }
            ], max_tokens=400)
        except Exception:
            return ""

    return ""

# -- Chat Endpoint -------------------------------------------------------------
@app.post("/api/chat")
async def chat(req: ChatRequest):
    # ── Deteksi #laporan ──────────────────────────────────────────────────────
    LAPORAN_TRIGGERS = ["#laporan","#Laporan", "#report", "#lpr"]
    matched_laporan  = None
    for trigger in LAPORAN_TRIGGERS:
        if req.message.lower().startswith(trigger):
            matched_laporan = trigger
            break

    if matched_laporan:
        laporan_text = req.message[len(matched_laporan):].strip()
        if not laporan_text:
            return {
                "type": "laporan_info",
                "message": (
                    "📋 Format pengiriman laporan:\n\n"
                    "#laporan [isi laporan]\n\n"
                    "Contoh:\n"
                    "#laporan Pekerjaan Rotating MA7 26 Mei 2026\n"
                    "Corrective Maintenance\n"
                    "1. 101-P-103: Perbaikan koneksi SAF (done)"
                ),
                "data": None, "columns": [], "chart": None,
                "narrative": None, "sql": None, "row_count": 0
            }

        items  = parse_laporan_with_ai(laporan_text)
        if not items:
            return {
                "type": "laporan_error",
                "message": (
                    "⚠️ Gagal memparse laporan.\n\n"
                    "Pastikan ada: tanggal, disiplin, dan daftar pekerjaan."
                ),
                "data": None, "columns": [], "chart": None,
                "narrative": None, "sql": None, "row_count": 0
            }

        success_count, error = insert_daily_report(items, req.pengirim, laporan_text)
        if error:
            return {
                "type": "laporan_error",
                "message": f"⚠️ Gagal menyimpan laporan: {error}",
                "data": None, "columns": [], "chart": None,
                "narrative": None, "sql": None, "row_count": 0
            }

        summary = {}
        for item in items[:success_count]:
            key = f"{item.get('disiplin', '-')} - {item.get('kategori', '-')}"
            summary[key] = summary.get(key, 0) + 1
        summary_lines = "\n".join([f"  • {k}: {v} item" for k, v in summary.items()])

        return {
            "type": "laporan_success",
            "message": (
                f"✅ Laporan berhasil disimpan!\n\n"
                f"📋 Total: {success_count} kegiatan tercatat\n\n"
                f"Rincian:\n{summary_lines}"
            ),
            "data": items[:success_count],
            "columns": ["tanggal_laporan", "disiplin", "direksi", "kategori", "tag_number", "deskripsi", "status_pekerjaan"],
            "chart": None,
            "narrative": None,
            "sql": None,
            "row_count": success_count
        }

    # ── Normal chat flow ──────────────────────────────────────────────────────
    dynamic_context = smart_entity_search(req.message)
    system_prompt   = BASE_SYSTEM_PROMPT + dynamic_context

    messages = [{"role": "system", "content": system_prompt}]
    for h in req.history[-10:]:
        messages.append({"role": h["role"], "content": h["content"]})
    messages.append({"role": "user", "content": req.message})

    try:
        raw = call_ai(messages, max_tokens=1500)

        json_match = re.search(r'\{[\s\S]*\}', raw)
        if not json_match:
            return {
                "type": "narrative", "message": raw, "data": None,
                "columns": [], "chart": None, "narrative": raw,
                "sql": None, "row_count": 0
            }

        parsed        = json.loads(json_match.group())
        response_type = parsed.get("type", "narrative")

        if response_type == "clarification":
            return {
                "type": "clarification",
                "message": parsed.get("clarification_question", parsed.get("message", "")),
                "data": None, "columns": [], "chart": None,
                "narrative": None, "sql": None, "row_count": 0
            }

        if response_type == "error":
            return {
                "type": "error", "message": parsed.get("message", "Terjadi kesalahan."),
                "data": None, "columns": [], "chart": None,
                "narrative": None, "sql": None, "row_count": 0
            }

        if response_type == "query" and parsed.get("sql"):
            sql          = parsed["sql"]
            valid, val_result = validate_sql(sql)
            if not valid:
                return {
                    "type": "error", "message": val_result, "data": None,
                    "columns": [], "chart": None, "narrative": None,
                    "sql": None, "row_count": 0
                }

            data, columns = execute_query(val_result)
            row_count     = len(data)

            narrative = ""
            if 0 < row_count <= 5 and len(columns) <= 5:
                narrative = generate_narrative(data, columns, req.message, parsed.get("narrative_hint", ""))

            chart = None
            if parsed.get("chart_suggestion") and row_count > 0:
                chart = {"type": parsed["chart_suggestion"], "config": parsed.get("chart_config", {})}

            serializable_data = []
            for row in data:
                clean_row = {}
                for k, v in row.items():
                    clean_row[k] = v.isoformat() if hasattr(v, 'isoformat') else v
                serializable_data.append(clean_row)

            return {
                "type": "query", "message": parsed.get("explanation", ""),
                "data": serializable_data, "columns": columns, "chart": chart,
                "narrative": narrative, "sql": val_result, "row_count": row_count
            }

        return {
            "type": "narrative", "message": parsed.get("message", raw),
            "data": None, "columns": [], "chart": None,
            "narrative": parsed.get("message", raw), "sql": None, "row_count": 0
        }

    except json.JSONDecodeError:
        return {
            "type": "narrative", "message": raw, "data": None,
            "columns": [], "chart": None, "narrative": raw,
            "sql": None, "row_count": 0
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# -- Laporan Endpoint (dedicated) ----------------------------------------------
@app.post("/api/laporan")
async def submit_laporan(req: LaporanRequest):
    """Endpoint khusus untuk submit laporan harian dari web UI."""
    if not req.raw_text.strip():
        raise HTTPException(status_code=400, detail="Teks laporan tidak boleh kosong")

    items = parse_laporan_with_ai(req.raw_text)
    if not items:
        raise HTTPException(status_code=422, detail="Gagal memparse laporan. Pastikan ada tanggal, disiplin, dan daftar pekerjaan.")

    success_count, error = insert_daily_report(items, req.pengirim, req.raw_text)
    if error:
        raise HTTPException(status_code=500, detail=f"Gagal menyimpan: {error}")

    return {
        "success": True,
        "total_saved": success_count,
        "items": items[:success_count]
    }

# -- Download Excel ------------------------------------------------------------
@app.post("/api/download")
async def download_excel(req: DownloadRequest):
    data, columns = execute_query(req.sql)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Export"

    header_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    alt_fill = PatternFill(start_color="E8EEF8", end_color="E8EEF8", fill_type="solid")
    for row_idx, row in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val  = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{req.filename}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/health")
def health():
    return {"status": "ok", "message": "Refinery Contract Chatbot API running"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)